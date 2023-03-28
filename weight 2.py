import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('example.xlsx')

# Select the active sheet
sheet = wb.active

# Loop through each row in the sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Extract the number from the first column
    number = str(row[0])
    
    # Shuffle the first and last digits
    if len(number) > 1:
        first_digit = number[0]
        last_digit = number[-1]
        middle_digits = number[1:-1]
        new_number = last_digit + middle_digits + first_digit
    else:
        new_number = number
    
    # Update the cell with the shuffled number
    sheet.cell(row=row[0].row, column=1, value=new_number)

# Save the updated Excel file
wb.save('example_shuffled.xlsx')

